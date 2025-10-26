"""
Excel -> ReqIF bridge

Opens a Windows file picker to select an Excel file, reads the "Requirements"
and "Relations" sheets, and generates a ReqIF using generate_reqif.ReqIFGenerator.

Expected columns in Requirements sheet:
- IE PUID (required)
- Type (required): functional | interface | performance
- ForeignID (required): integer-like
- Name (required)
- Chapter (optional)
- Description (optional)
- Text (recommended): plain text; newline-separated paragraphs; lines starting with "- " become list items
- TextHTML (optional): ignored unless you want to extend parsing; plain Text is preferred
- Status (optional): draft | wip | reviewed | approved
- Priority (optional): high | medium | low
- ReqPrefix (optional)
- Identifier (optional)
- Order (optional): integer for ordering

Expected columns in Relations sheet:
- RelationType (required): satisfy | derive | refine
- SourceIEPUID (required): IE PUID of source requirement
- TargetIEPUID (required): IE PUID of target requirement
- Identifier (optional)
"""

import os
import sys
from typing import List, Dict, Optional

# UI file dialog
try:
    import tkinter as tk
    from tkinter import filedialog, messagebox
except Exception:
    tk = None
    filedialog = None
    messagebox = None

# Excel loaders
_HAVE_PANDAS = False
_HAVE_OPENPYXL = False
try:
    import pandas as pd  # type: ignore
    _HAVE_PANDAS = True
except Exception:
    try:
        import openpyxl  # type: ignore
        _HAVE_OPENPYXL = True
    except Exception:
        pass

from xml.etree.ElementTree import Element, SubElement

# Import the generator, working both when run as a module and as a script
try:
    from reqif_app.generate_reqif import ReqIFGenerator  # when installed/run as package
except Exception:  # pragma: no cover
    try:
        from generate_reqif import ReqIFGenerator  # when run from same folder
    except Exception:
        import os, sys
        sys.path.append(os.path.dirname(__file__))
        from generate_reqif import ReqIFGenerator


def build_xhtml_from_text(text: str) -> Element:
    """Convert simple plain text to XHTML: paragraphs and unordered lists.

    - Blank lines separate paragraphs
    - Consecutive lines starting with "- " become a single <ul> with <li>
    """
    ns = '{http://www.w3.org/1999/xhtml}'
    div = Element(f'{ns}div')
    if not text:
        SubElement(div, f'{ns}p').text = ''
        return div

    lines = text.replace('\r\n', '\n').replace('\r', '\n').split('\n')
    buffer: List[str] = []
    list_mode = False
    ul_elem: Optional[Element] = None

    def flush_paragraph(buf: List[str]):
        content = '\n'.join(buf).strip()
        if content:
            SubElement(div, f'{ns}p').text = content

    for raw in lines:
        line = raw.rstrip()
        if not line:
            # blank line: end paragraph or list
            if list_mode:
                list_mode = False
                ul_elem = None
            elif buffer:
                flush_paragraph(buffer)
                buffer = []
            continue

        if line.startswith('- '):
            # entering list mode
            if not list_mode:
                if buffer:
                    flush_paragraph(buffer)
                    buffer = []
                ul_elem = SubElement(div, f'{ns}ul')
                list_mode = True
            li = SubElement(ul_elem, f'{ns}li')  # type: ignore[arg-type]
            li.text = line[2:].strip()
        else:
            # normal paragraph text
            if list_mode:
                list_mode = False
                ul_elem = None
            buffer.append(line)

    # trailing content
    if buffer:
        flush_paragraph(buffer)

    return div


def _ensure_int(val, default: int = 0) -> int:
    try:
        if val is None or (isinstance(val, float) and pd.isna(val)):  # type: ignore[name-defined]
            return default
    except Exception:
        if val is None:
            return default
    try:
        return int(val)
    except Exception:
        try:
            return int(str(val).strip())
        except Exception:
            return default


def _normalize_key(s: str) -> str:
    s = str(s).strip().lower()
    # replace spaces and hyphens with underscore
    for ch in [' ', '-', '\t']:
        s = s.replace(ch, '_')
    # collapse double underscores
    while '__' in s:
        s = s.replace('__', '_')
    return s


def _normalize_records(records: List[Dict[str, object]]) -> List[Dict[str, object]]:
    out: List[Dict[str, object]] = []
    for rec in records:
        nrec: Dict[str, object] = {}
        for k, v in rec.items():
            nrec[_normalize_key(k)] = v
        out.append(nrec)
    return out


def _read_excel_with_pandas(path: str) -> Dict[str, List[Dict[str, object]]]:
    frames = pd.read_excel(path, sheet_name=['Requirements', 'Relations'])  # type: ignore[name-defined]
    out: Dict[str, List[Dict[str, object]]] = {}
    for name, df in frames.items():
        df = df.rename(columns={c: str(c).strip() for c in df.columns})
        recs = df.fillna('').to_dict(orient='records')
        out[name] = _normalize_records(recs)
    return out


def _read_excel_with_openpyxl(path: str) -> Dict[str, List[Dict[str, object]]]:
    from openpyxl import load_workbook  # type: ignore
    wb = load_workbook(path, data_only=True)
    required_sheets = ['Requirements', 'Relations']
    for s in required_sheets:
        if s not in wb.sheetnames:
            raise ValueError(f"Missing sheet: {s}")

    def sheet_to_dicts(ws) -> List[Dict[str, object]]:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else '' for h in rows[0]]
        records: List[Dict[str, object]] = []
        for r in rows[1:]:
            rec = {headers[i]: (r[i] if i < len(r) else '') for i in range(len(headers))}
            records.append(rec)
        return _normalize_records(records)

    return {
        'Requirements': sheet_to_dicts(wb['Requirements']),
        'Relations': sheet_to_dicts(wb['Relations']),
    }


def pick_excel_file(initial: Optional[str] = None) -> Optional[str]:
    if tk is None or filedialog is None:
        return initial if initial and os.path.isfile(initial) else None
    root = tk.Tk()
    root.withdraw()
    path = filedialog.askopenfilename(
        title='Select requirements Excel file',
        initialdir=os.path.dirname(initial) if initial else None,
        filetypes=[('Excel files', '*.xlsx;*.xlsm;*.xltx;*.xltm'), ('All files', '*.*')]
    )
    root.update()
    root.destroy()
    return path or None


def main():
    default_path = r"C:\\Users\\Asus\\OneDrive\\Work\\ReqIF_Generator\\requirements_template.xlsx"
    excel_path = pick_excel_file(default_path)
    if not excel_path:
        if messagebox:
            messagebox.showinfo('ReqIF From Excel', 'No file selected.')
        else:
            print('No file selected.')
        sys.exit(1)
    if not os.path.isfile(excel_path):
        if messagebox:
            messagebox.showerror('ReqIF From Excel', f'File not found:\n{excel_path}')
        else:
            print(f'File not found: {excel_path}')
        sys.exit(1)

    # Load Excel data
    try:
        if _HAVE_PANDAS:
            data = _read_excel_with_pandas(excel_path)
        elif _HAVE_OPENPYXL:
            data = _read_excel_with_openpyxl(excel_path)
        else:
            msg = 'Neither pandas nor openpyxl is available. Please install one of them to read Excel files.'
            if messagebox:
                messagebox.showerror('ReqIF From Excel', msg)
            else:
                print(msg)
            sys.exit(2)
    except Exception as e:
        if messagebox:
            messagebox.showerror('ReqIF From Excel', f'Failed to read Excel:\n{e}')
        else:
            print(f'Failed to read Excel: {e}')
        sys.exit(2)

    req_rows = data.get('Requirements', [])
    rel_rows = data.get('Relations', [])

    warnings: List[str] = []

    # Initialize generator
    generator = ReqIFGenerator(title='System Requirements Specification')

    # Add requirements (optionally sorted by Order)
    def order_key(rec):
        v = rec.get('order', '')
        try:
            return int(v)
        except Exception:
            return 1_000_000

    req_rows_sorted = sorted(req_rows, key=order_key)

    # Map IE PUID -> SPEC-OBJECT identifier
    id_map: Dict[str, str] = {}

    TYPE_MAP = {
        'functional': 'functional', 'func': 'functional', 'fr': 'functional', 'f': 'functional',
        'interface': 'interface', 'if': 'interface', 'ir': 'interface', 'i': 'interface',
        'performance': 'performance', 'perf': 'performance', 'pr': 'performance', 'p': 'performance',
        'non-functional': 'performance', 'nfr': 'performance'
    }

    added_reqs = 0

    for row in req_rows_sorted:
        raw_type = str(row.get('type', row.get('req_type', ''))).strip().lower()
        type_val = TYPE_MAP.get(raw_type)
        if not type_val:
            # Try to normalize phrases like "functional requirement"
            if 'functional' in raw_type:
                type_val = 'functional'
            elif 'interface' in raw_type:
                type_val = 'interface'
            elif 'performance' in raw_type or 'non-functional' in raw_type:
                type_val = 'performance'
            else:
                warnings.append(f"Row skipped: invalid Type '{row.get('type') or row.get('req_type')}'.")
                continue

        ie_puid = str(row.get('ie_puid', row.get('iepuid', ''))).strip()
        foreign_id = _ensure_int(row.get('foreign_id', row.get('foreignid', '')))
        name = str(row.get('name', '')).strip()
        chapter = str(row.get('chapter', '')).strip()
        description = str(row.get('description', row.get('desc', ''))).strip()
        text_plain = str(row.get('text', row.get('text_content', '')) or '').strip()
        text_html = str(row.get('texthtml', row.get('text_html', '')) or '').strip()
        status = str(row.get('status', '') or '').strip() or 'approved'
        priority = str(row.get('priority', '') or '').strip() or 'medium'
        req_prefix = str(row.get('reqprefix', row.get('req_prefix', '')) or '').strip() or {
            'functional': 'SYS-F', 'interface': 'SYS-I', 'performance': 'SYS-P'
        }[type_val]
        identifier = str(row.get('identifier', row.get('spec_object_id', '')) or '').strip() or None

        # Build text content (prefer plain text -> XHTML); ignore TextHTML unless you extend parsing
        if text_plain:
            text_content = build_xhtml_from_text(text_plain)
        elif text_html and text_html.startswith('<'):
            # Fallback: wrap raw markup as plain text paragraph to avoid breaking XML
            text_content = build_xhtml_from_text(text_html)
        else:
            text_content = build_xhtml_from_text(description or name)

        so_id = generator.add_requirement(
            req_type=type_val,
            foreign_id=foreign_id,
            name=name or ie_puid or 'Untitled',
            chapter=chapter,
            description=description,
            text_content=text_content,
            status=status,
            priority=priority,
            req_prefix=req_prefix,
            identifier=identifier,
            ie_puid=ie_puid if ie_puid else None,
        )

        if ie_puid:
            id_map[ie_puid] = so_id
        added_reqs += 1

    # Add relations
    added_rels = 0
    for row in rel_rows:
        rtype = str(row.get('relationtype', row.get('relation_type', ''))).strip().lower()
        # Accept either mapping by IE PUID or by SPEC-OBJECT identifier
        src_ie = str(row.get('sourceiepuid', row.get('source_ie_puid', ''))).strip()
        tgt_ie = str(row.get('targetiepuid', row.get('target_ie_puid', ''))).strip()
        src_id_direct = str(row.get('source_id', '')).strip()
        tgt_id_direct = str(row.get('target_id', '')).strip()
        rid = str(row.get('identifier', '') or '').strip() or None

        if rtype not in {'satisfy', 'derive', 'refine'}:
            warnings.append(f"Relation skipped: invalid RelationType '{row.get('RelationType')}'.")
            continue
        if src_id_direct and tgt_id_direct:
            generator.add_relation(rtype, source_id=src_id_direct, target_id=tgt_id_direct, identifier=rid)
            added_rels += 1
            continue
        if not src_ie or not tgt_ie:
            warnings.append('Relation skipped: missing SourceIEPUID/TargetIEPUID and no direct IDs provided.')
            continue
        src = id_map.get(src_ie)
        tgt = id_map.get(tgt_ie)
        if not src or not tgt:
            warnings.append(f"Relation skipped: could not resolve IDs: {src_ie} -> {tgt_ie}.")
            continue
        generator.add_relation(rtype, source_id=src, target_id=tgt, identifier=rid)
        added_rels += 1

    # Output
    out_name = os.path.splitext(os.path.basename(excel_path))[0] + '.reqif'
    out_path = os.path.join(os.path.dirname(excel_path), out_name)
    generator.generate(out_path)

    summary = [
        f'Generated: {out_path}',
        f'Requirements added: {added_reqs}',
        f'Relations added: {added_rels}'
    ]
    if warnings:
        summary.append(f'Warnings: {len(warnings)} (showing up to 10)')
        summary.extend([f' - {w}' for w in warnings[:10]])
    message = '\n'.join(summary)
    if messagebox:
        messagebox.showinfo('ReqIF From Excel', message)
    else:
        print('\n' + message)


if __name__ == '__main__':
    main()
